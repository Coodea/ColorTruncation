#!/usr/bin/env python3
# binary_bitmap_viewer.py — scrollable + zoomable bitmap/diff for huge binary CSVs
# Row-accurate view + Save PNG + Save CSV (0/1) + Save XLSX with red boxes.
#
# CSV input:
#   Header R,G,B (tolerant)
#   Rows: 8-bit strings (10101100,00011010,11110000) OR decimal 0..255 -> auto 8-bit
#
# View:
#   • Single (Before/After) or XOR Diff
#   • Max rows (row limit), Manual stride (row sampling)
#   • Adjustable cell size
#   • Zoom 10–400%, scrollable canvas
#   • Save PNG …, Save CSV of view … (0/1), Save XLSX of view … (red boxes)
#
# Mouse: Wheel = vertical scroll | Shift+Wheel = horizontal | Ctrl+Wheel = zoom

import csv
import tkinter as tk
from tkinter import filedialog, messagebox
from typing import List, Tuple
import numpy as np
from PIL import Image, ImageTk, ImageDraw

# -------------------- CSV parsing --------------------

def _clean_cell(x: str) -> str:
    x = (x or "").strip()
    if len(x) >= 2 and ((x[0] == x[-1] == '"') or (x[0] == x[-1] == "'")):
        x = x[1:-1]
    return x.strip()

def _to_8bit_binary(cell: str) -> str:
    c = _clean_cell(cell)
    if c.lower().startswith("0b"):
        c = c[2:]
    if len(c) == 8 and set(c) <= {"0","1"}:
        return c
    try:
        v = int(c)
        if 0 <= v <= 255:
            return f"{v:08b}"
    except Exception:
        pass
    return ""

def _merge_rgb_bits(r: str, g: str, b: str) -> str:
    rb = _to_8bit_binary(r); gb = _to_8bit_binary(g); bb = _to_8bit_binary(b)
    return (rb + gb + bb) if (rb and gb and bb) else ""

def read_binary_csv_bits(path: str) -> Tuple[np.ndarray, int, int]:
    """
    Returns:
      bitmat: (N, 24) boolean for bits [R(8)|G(8)|B(8)] left-to-right
      parsed_rows: total rows parsed from the CSV
      ones_count: number of '1' bits across all parsed rows
    """
    rows_bits: List[str] = []
    parsed_rows = 0
    ones_count = 0
    with open(path, "r", encoding="utf-8", newline="") as f:
        rdr = csv.reader(f)
        first = True
        for row in rdr:
            if not row or len(row) < 3:
                continue
            if first:
                head = ",".join(_clean_cell(x).lower() for x in row[:3])
                if head in ("r,g,b","r, g, b"):
                    first = False
                    continue
                first = False
            s24 = _merge_rgb_bits(row[0], row[1], row[2])
            if not s24:
                continue
            rows_bits.append(s24)
            parsed_rows += 1
            ones_count += s24.count("1")

    if not rows_bits:
        return np.zeros((0,24), bool), 0, 0

    arr = np.frombuffer("".join(rows_bits).encode("ascii"), dtype="S1").reshape(-1,24)
    return (arr == b"1"), parsed_rows, ones_count

# -------------------- Rendering (row-accurate) --------------------

def select_rows(bitmat: np.ndarray, max_rows: int, stride: int) -> np.ndarray:
    """
    Row-accurate selection (no aggregation).
    - Takes the first 'max_rows' rows (if 0 or None -> all)
    - Then samples every 'stride' rows (stride >= 1)
    """
    if bitmat.size == 0:
        return bitmat
    n = bitmat.shape[0]
    mrows = max_rows if (max_rows and max_rows > 0) else n
    mrows = min(n, mrows)
    s = max(1, int(stride))
    return bitmat[:mrows:s, :]

def make_grid_image(grid: np.ndarray, cell: int = 10, pad: int = 5,
                    draw_grid: bool = True) -> Image.Image:
    """
    grid: (rows, 24) boolean. Draw as red where True.
    Uses NEAREST scaling to keep pixels crisp.
    """
    rows, cols = grid.shape
    cols = max(cols, 24)
    cell = max(2, int(cell))
    pad  = max(2, int(pad))

    W = pad*2 + cols*cell
    H = pad*2 + rows*cell
    img = Image.new("RGB", (W, H), "white")

    # paint: tiny mask -> upscale
    if rows > 0:
        mask = np.zeros((rows, cols), dtype=np.uint8)
        mask[grid] = 255
        mask_img = Image.fromarray(mask, mode="L").resize((cols*cell, rows*cell), Image.NEAREST)
        img.paste(Image.new("RGB", mask_img.size, (200,0,0)), (pad, pad), mask_img)

    if draw_grid:
        draw = ImageDraw.Draw(img)
        # light grid
        for c in range(cols+1):
            x = pad + c*cell
            draw.line([(x, pad), (x, H-pad)], fill=(200,200,200), width=1)
        for r in range(rows+1):
            y = pad + r*cell
            draw.line([(pad, y), (W-pad, y)], fill=(220,220,220), width=1)
        # thick red separators after 8 & 16
        for sep in (8,16):
            x = pad + sep*cell
            draw.line([(x, pad), (x, H-pad)], fill=(200,0,0), width=3)
        # labels
        for idx, name in zip((4,12,20), ("R","G","B")):
            draw.text((pad + idx*cell - 4, 4), name, fill=(120,120,120))

    return img

def render_single(bitmat: np.ndarray, max_rows: int, stride: int,
                  cell: int = 10):
    """
    Returns (image, rows_loaded, rows_shown, grid_used)
    """
    n = bitmat.shape[0]
    if n == 0:
        return Image.new("RGB", (320,120), "white"), 0, 0, np.zeros((0,24), dtype=bool)
    sel = select_rows(bitmat, max_rows=max_rows, stride=stride)
    img = make_grid_image(sel, cell=cell, pad=max(2, cell//2), draw_grid=True)
    return img, n, sel.shape[0], sel

def render_diff(before_bits: np.ndarray, after_bits: np.ndarray,
                max_rows: int, stride: int, cell: int = 10):
    """
    XOR per-row (no aggregation). Uses the first min(n_before, n_after) rows.
    Then applies max_rows/stride selection to the XOR matrix.
    Returns (image, rows_loaded=min(n_b,n_a), rows_shown, grid_used)
    """
    n = min(before_bits.shape[0], after_bits.shape[0])
    if n == 0:
        return Image.new("RGB", (320,120), "white"), 0, 0, np.zeros((0,24), dtype=bool)
    diff = before_bits[:n] ^ after_bits[:n]
    sel = select_rows(diff, max_rows=max_rows, stride=stride)
    img = make_grid_image(sel, cell=cell, pad=max(2, cell//2), draw_grid=True)
    return img, n, sel.shape[0], sel

# -------------------- GUI --------------------

class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Binary CSV → Bit Map (Diff) • Scroll & Zoom • Row-accurate")
        self.root.geometry("1100x800")

        self.before_path = tk.StringVar()
        self.after_path  = tk.StringVar()

        # view params
        self.cell_size   = tk.IntVar(value=12)     # base draw cell size
        self.max_rows    = tk.IntVar(value=100000) # hard limit of rows to display (row-accurate)
        self.stride      = tk.IntVar(value=1)      # sampling step (1 = every row)
        self.zoom_pct    = tk.IntVar(value=100)    # 10..400% display zoom

        # storage of the last rendered grid (for saving)
        self.last_grid = None  # np.ndarray (rows,24) of bool
        self.last_mode = "single"  # or "diff"

        # Top file pickers
        top = tk.Frame(self.root); top.pack(fill="x", padx=10, pady=(10,6))
        tk.Label(top, text="Before (binary CSV):").grid(row=0, column=0, sticky="w")
        tk.Entry(top, textvariable=self.before_path, width=78).grid(row=0, column=1, sticky="we", padx=6)
        tk.Button(top, text="Browse…", command=self.browse_before).grid(row=0, column=2)

        tk.Label(top, text="After (binary CSV):").grid(row=1, column=0, sticky="w", pady=(6,0))
        tk.Entry(top, textvariable=self.after_path, width=78).grid(row=1, column=1, sticky="we", padx=6, pady=(6,0))
        tk.Button(top, text="Browse…", command=self.browse_after).grid(row=1, column=2, pady=(6,0))

        # Options
        opts = tk.Frame(self.root); opts.pack(fill="x", padx=10, pady=6)
        tk.Label(opts, text="Cell size").pack(side="left")
        tk.Spinbox(opts, textvariable=self.cell_size, from_=2, to=40, width=5).pack(side="left", padx=(4,12))

        tk.Label(opts, text="Max rows").pack(side="left")
        tk.Spinbox(opts, textvariable=self.max_rows, from_=100, to=5_000_000, increment=1000, width=10)\
            .pack(side="left", padx=(4,12))

        tk.Label(opts, text="Manual stride").pack(side="left")
        tk.Spinbox(opts, textvariable=self.stride, from_=1, to=5000, width=7).pack(side="left", padx=(4,12))

        tk.Button(opts, text="Render", command=self.render).pack(side="left", padx=8)
        tk.Button(opts, text="Save PNG…", command=self.save_png).pack(side="left", padx=(0,8))
        tk.Button(opts, text="Save CSV of view…", command=self.save_csv_view).pack(side="left", padx=(0,8))
        tk.Button(opts, text="Save XLSX of view…", command=self.save_xlsx_view).pack(side="left", padx=(0,8))
        tk.Button(opts, text="Quit", command=self.root.destroy).pack(side="right")

        # Zoom controls
        zoomf = tk.Frame(self.root); zoomf.pack(fill="x", padx=10, pady=(0,6))
        tk.Label(zoomf, text="Zoom").pack(side="left")
        tk.Scale(zoomf, variable=self.zoom_pct, from_=10, to=400, orient="horizontal", length=240,
                 command=lambda _=None: self.redraw_zoom()).pack(side="left", padx=(4,12))
        tk.Button(zoomf, text="Fit Width", command=self.fit_width).pack(side="left")

        # Status
        self.status = tk.StringVar(value="Choose CSVs and click Render.")
        tk.Label(self.root, textvariable=self.status, anchor="w").pack(fill="x", padx=10)

        # Scrollable canvas
        wrap = tk.Frame(self.root); wrap.pack(fill="both", expand=True, padx=10, pady=(0,10))
        self.vbar = tk.Scrollbar(wrap, orient="vertical")
        self.hbar = tk.Scrollbar(wrap, orient="horizontal")
        self.canvas = tk.Canvas(wrap, bg="white",
                                xscrollcommand=self.hbar.set,
                                yscrollcommand=self.vbar.set)
        self.vbar.config(command=self.canvas.yview)
        self.hbar.config(command=self.canvas.xview)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.vbar.grid(row=0, column=1, sticky="ns")
        self.hbar.grid(row=1, column=0, sticky="we")
        wrap.rowconfigure(0, weight=1); wrap.columnconfigure(0, weight=1)

        # Mouse wheel bindings
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind("<Shift-MouseWheel>", self._on_mousewheel_h)
        self.canvas.bind("<Control-MouseWheel>", self._on_mousewheel_zoom)

        self.tk_img = None
        self.base_img = None
        self.display_img = None
        self.root.mainloop()

    # ----- File pickers -----
    def browse_before(self):
        p = filedialog.askopenfilename(title="Choose BEFORE binary CSV",
                                       filetypes=[("CSV files","*.csv"), ("All files","*.*")])
        if p: self.before_path.set(p)

    def browse_after(self):
        p = filedialog.askopenfilename(title="Choose AFTER binary CSV",
                                       filetypes=[("CSV files","*.csv"), ("All files","*.*")])
        if p: self.after_path.set(p)

    # ----- Mouse wheel handlers -----
    def _on_mousewheel(self, event):
        delta = -1 * (event.delta // 120 if event.delta else 0)
        self.canvas.yview_scroll(delta, "units")

    def _on_mousewheel_h(self, event):
        delta = -1 * (event.delta // 120 if event.delta else 0)
        self.canvas.xview_scroll(delta, "units")

    def _on_mousewheel_zoom(self, event):
        d = 1 if event.delta > 0 else -1
        z = self.zoom_pct.get()
        z = max(10, min(400, z + d*10))
        self.zoom_pct.set(z)
        self.redraw_zoom()

    # ----- Rendering -----
    def render(self):
        b = self.before_path.get().strip()
        a = self.after_path.get().strip()
        if not b and not a:
            messagebox.showwarning("No input", "Select at least one CSV.")
            return
        try:
            cell   = self.cell_size.get()
            maxr   = self.max_rows.get()
            stride = self.stride.get()

            if b and a:
                bits_b, rows_b, ones_b = read_binary_csv_bits(b)
                bits_a, rows_a, ones_a = read_binary_csv_bits(a)
                if bits_b.size == 0 or bits_a.size == 0:
                    messagebox.showerror("No data", "One of the CSVs has no usable rows.")
                    return
                img, rows_loaded, rows_shown, grid = render_diff(
                    bits_b, bits_a,
                    max_rows=maxr,
                    stride=stride,
                    cell=cell
                )
                self.last_grid = grid
                self.last_mode = "diff"
                self.status.set(
                    f"Before parsed: {rows_b} rows | After parsed: {rows_a} rows "
                    f"| Compared rows (min): {rows_loaded} | Displayed: {rows_shown} "
                    f"(max_rows={maxr}, stride={stride}, cell={cell})"
                )
            else:
                path = b or a
                bits, rows, ones = read_binary_csv_bits(path)
                if bits.size == 0:
                    messagebox.showerror("No data", "The CSV appears empty or not in expected format.")
                    return
                img, rows_loaded, rows_shown, grid = render_single(
                    bits,
                    max_rows=maxr,
                    stride=stride,
                    cell=cell
                )
                self.last_grid = grid
                self.last_mode = "single"
                which = "Before" if b else "After"
                self.status.set(
                    f"{which} parsed: {rows} rows | Displayed: {rows_shown} "
                    f"(max_rows={maxr}, stride={stride}, cell={cell})"
                )

            self.base_img = img
            self.redraw_zoom()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def redraw_zoom(self):
        if self.base_img is None:
            return
        z = self.zoom_pct.get() / 100.0
        w, h = self.base_img.size
        zw = max(1, int(round(w * z)))
        zh = max(1, int(round(h * z)))
        self.display_img = self.base_img if (zw == w and zh == h) else self.base_img.resize((zw, zh), Image.NEAREST)

        self.tk_img = ImageTk.PhotoImage(self.display_img)
        self.canvas.delete("all")
        self.canvas.create_image(0, 0, image=self.tk_img, anchor="nw")
        self.canvas.config(scrollregion=(0, 0, self.display_img.size[0], self.display_img.size[1]))

    def fit_width(self):
        if self.base_img is None:
            return
        cw = max(50, self.canvas.winfo_width())
        bw = self.base_img.size[0]
        z = max(10, min(400, int(round(cw / bw * 100))))
        self.zoom_pct.set(z)
        self.redraw_zoom()

    # ----- Save: PNG / CSV (0/1) / XLSX (red boxes) -----
    def save_png(self):
        if self.base_img is None:
            messagebox.showinfo("Nothing to save", "Render first, then save.")
            return
        p = filedialog.asksaveasfilename(title="Save bitmap as PNG",
                                         defaultextension=".png",
                                         filetypes=[("PNG","*.png")])
        if not p:
            return
        try:
            self.base_img.save(p, "PNG")
            messagebox.showinfo("Saved", f"Saved to:\n{p}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def save_csv_view(self):
        if self.last_grid is None or self.last_grid.size == 0:
            messagebox.showinfo("Nothing to save", "Render first, then save.")
            return
        p = filedialog.asksaveasfilename(title="Save CSV of current view (0/1)",
                                         defaultextension=".csv",
                                         filetypes=[("CSV files","*.csv")])
        if not p:
            return
        try:
            rows, cols = self.last_grid.shape
            hdr1 = ["R"]*8 + ["G"]*8 + ["B"]*8
            hdr2 = [f"b{b}" for b in range(7,-1,-1)]*3
            with open(p, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                w.writerow(hdr1); w.writerow(hdr2)
                for r in range(rows):
                    w.writerow(["1" if x else "0" for x in self.last_grid[r].tolist()])
            messagebox.showinfo("Saved", f"Saved CSV to:\n{p}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def save_xlsx_view(self):
        """
        Export the current grid as an Excel file with red-filled cells for True bits.
        This preserves the 'red boxes' look.
        """
        if self.last_grid is None or self.last_grid.size == 0:
            messagebox.showinfo("Nothing to save", "Render first, then save.")
            return
        p = filedialog.asksaveasfilename(title="Save XLSX of current view (red boxes)",
                                         defaultextension=".xlsx",
                                         filetypes=[("Excel Workbook","*.xlsx")])
        if not p:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Alignment, Font
            wb = Workbook()
            ws = wb.active
            ws.title = "BitMap"

            rows, cols = self.last_grid.shape

            # Headers (two rows)
            hdr1 = ["R"]*8 + ["G"]*8 + ["B"]*8
            hdr2 = [f"b{b}" for b in range(7,-1,-1)]*3
            ws.append(hdr1)
            ws.append(hdr2)

            # Style headers
            hdr_font = Font(bold=True)
            for c in range(1, cols+1):
                ws.cell(1, c).font = hdr_font
                ws.cell(2, c).font = hdr_font
                ws.cell(1, c).alignment = Alignment(horizontal="center")
                ws.cell(2, c).alignment = Alignment(horizontal="center")

            # Fills (reuse objects)
            red_fill = PatternFill(fill_type="solid", fgColor="FFC80000")  # opaque red
            # Optional: white fill not necessary; leave defaults for zeros

            # Write grid (start at row 3)
            # For speed, only set fill on True cells.
            for r in range(rows):
                excel_row = r + 3
                # create empty row quickly
                for c in range(1, cols+1):
                    ws.cell(excel_row, c, "")  # keep cells blank
                # apply red fill where True
                trues = np.where(self.last_grid[r])[0]
                for j in trues:
                    cell = ws.cell(excel_row, j+1)
                    cell.fill = red_fill

                # Make rows short to look like squares; adjust as you like
                ws.row_dimensions[excel_row].height = 10  # points

            # Column widths to look square-ish
            for c in range(1, cols+1):
                ws.column_dimensions[chr(64+c) if c <= 26 else f"A{chr(64+c-26)}"].width = 2.0  # narrow columns

            wb.save(p)
            messagebox.showinfo("Saved", f"Saved XLSX to:\n{p}\n(Note: very large exports can be big/slow in Excel.)")
        except ImportError:
            messagebox.showerror("Missing library",
                                 "openpyxl is required for XLSX export.\nInstall with:\n\npip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", str(e))

if __name__ == "__main__":
    App()
